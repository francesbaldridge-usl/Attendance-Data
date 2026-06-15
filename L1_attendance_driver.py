# import pandas as pd
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from selenium.common.exceptions import TimeoutException, NoSuchElementException
# from selenium.webdriver.edge.options import Options
# import time

# options = Options()
# options.add_argument("--headless")
# options.add_argument("--window-size=1920,1080")
# options.add_argument("--blink-settings=imagesEnabled=true")
# options.add_argument("--no-sandbox")
# options.add_argument("--disable-dev-shm-usage")

# URL = "https://www.uslleagueone.com/league-schedule"
# WAIT_SECONDS = 15


# def get_team_name(row, side):
#     """Try crest image alt first (full name), fall back to short code text."""
#     try:
#         img = row.find_element(By.CSS_SELECTOR, f"td.Opta-Crest.Opta-{side} img")
#         name = img.get_attribute("alt")
#         if name:
#             return name.strip()
#     except NoSuchElementException:
#         pass
#     try:
#         td = row.find_element(By.CSS_SELECTOR, f"td.Opta-Team.Opta-TeamName.Opta-{side}")
#         return td.get_attribute("innerHTML").strip() or None
#     except NoSuchElementException:
#         return None


# def scrape_schedule(url: str) -> pd.DataFrame:
#     driver = webdriver.Edge(options=options)
#     wait = WebDriverWait(driver, WAIT_SECONDS)
#     records = []

#     try:
#         driver.get(url)
#         wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "tbody.Opta-fixture")))
#         time.sleep(3)

#         total = len(driver.find_elements(By.CSS_SELECTOR, "table tbody"))
#         print(f"Total tbody rows found: {total}")

#         current_date = None

#         for i in range(total):
#             all_rows = driver.find_elements(By.CSS_SELECTOR, "table tbody")
#             if i >= len(all_rows):
#                 break
#             tbody = all_rows[i]

#             try:
#                 driver.execute_script("arguments[0].scrollIntoView({block:'center'});", tbody)
#                 time.sleep(0.5)
#                 classes = tbody.get_attribute("class") or ""
#             except Exception:
#                 continue

#             if "Opta-fixture" not in classes:
#                 try:
#                     span = tbody.find_element(By.CSS_SELECTOR, "tr > td > h4 > span")
#                     date_text = span.get_attribute('innerHTML').strip()
#                     if date_text:
#                         current_date = date_text
#                         print(f"\n  Date: {current_date}")
#                 except NoSuchElementException:
#                     pass
#                 continue

#             if "Opta-prematch" in classes:
#                 print(f"  [{i}] First unplayed match — stopping.")
#                 break

#             if "Opta-result" not in classes:
#                 continue

#             record = {"date": current_date, "home_team": None, "away_team": None, "attendance": None}

#             try:
#                 score_row = tbody.find_element(By.CSS_SELECTOR, "tr.Opta-Scoreline")
#                 record["home_team"] = get_team_name(score_row, "Home")
#                 record["away_team"] = get_team_name(score_row, "Away")

#                 button = score_row.find_element(By.CSS_SELECTOR, "button.Opta-Nest-Control")
#                 expansion_id = button.get_attribute("data-expansion_id")

#                 driver.execute_script("arguments[0].click();", button)

#                 try:
#                     wait.until(EC.presence_of_element_located((By.XPATH,
#                         f"//*[@id='{expansion_id}']//div[@class='Opta-Matchdata']//dt[text()='Attendance']"
#                     )))
#                     panel = driver.find_element(By.ID, expansion_id)
#                     att_el = panel.find_element(
#                         By.XPATH, ".//div[@class='Opta-Matchdata']//dt[text()='Attendance']/following-sibling::dd[1]"
#                     )
#                     record["attendance"] = att_el.get_attribute('innerHTML').strip() or None
#                 except (NoSuchElementException, TimeoutException):
#                     record["attendance"] = None

#                 try:
#                     all_rows2 = driver.find_elements(By.CSS_SELECTOR, "table tbody")
#                     tbody2 = all_rows2[i]
#                     score_row2 = tbody2.find_element(By.CSS_SELECTOR, "tr.Opta-Scoreline")
#                     button2 = score_row2.find_element(By.CSS_SELECTOR, "button.Opta-Nest-Control")
#                     driver.execute_script("arguments[0].click();", button2)
#                 except Exception:
#                     pass
#                 time.sleep(0.3)

#                 print(f"  [{i}] {record['date']} | {record['home_team']} vs {record['away_team']} | Att: {record['attendance']}")

#             except Exception as e:
#                 print(f"  [{i}] Error: {e}")

#             records.append(record)

#     finally:
#         driver.quit()

#     return pd.DataFrame(records)


# if __name__ == "__main__":
#     df = scrape_schedule(URL)
#     print(f"\n{df.head(10)}")
#     df.to_excel("L1_attendance_data.xlsx", index=False)

import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.edge.options import Options
from openpyxl import load_workbook
import time

options = Options()
options.add_argument("--headless")
options.add_argument("--window-size=1920,1080")
options.add_argument("--blink-settings=imagesEnabled=true")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")

URL = "https://www.uslleagueone.com/league-schedule"
WAIT_SECONDS = 15


def get_team_name(row, side):
    """Try crest image alt first (full name), fall back to short code text."""
    try:
        img = row.find_element(By.CSS_SELECTOR, f"td.Opta-Crest.Opta-{side} img")
        name = img.get_attribute("alt")
        if name:
            return name.strip()
    except NoSuchElementException:
        pass
    try:
        td = row.find_element(By.CSS_SELECTOR, f"td.Opta-Team.Opta-TeamName.Opta-{side}")
        return td.get_attribute("innerHTML").strip() or None
    except NoSuchElementException:
        return None


def scrape_schedule(url: str) -> pd.DataFrame:
    driver = webdriver.Edge(options=options)
    wait = WebDriverWait(driver, WAIT_SECONDS)
    records = []

    try:
        driver.get(url)
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "tbody.Opta-fixture")))
        time.sleep(3)

        total = len(driver.find_elements(By.CSS_SELECTOR, "table tbody"))
        print(f"Total tbody rows found: {total}")

        current_date = None

        for i in range(total):
            all_rows = driver.find_elements(By.CSS_SELECTOR, "table tbody")
            if i >= len(all_rows):
                break
            tbody = all_rows[i]

            try:
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", tbody)
                time.sleep(0.5)
                classes = tbody.get_attribute("class") or ""
            except Exception:
                continue

            if "Opta-fixture" not in classes:
                try:
                    span = tbody.find_element(By.CSS_SELECTOR, "tr > td > h4 > span")
                    date_text = span.get_attribute('innerHTML').strip()
                    if date_text:
                        try:
                            current_date = datetime.strptime(date_text, "%A, %B %d, %Y")
                        except ValueError:
                            try:
                                current_date = datetime.strptime(f"{date_text}, {datetime.now().year}", "%A, %B %d, %Y")
                            except ValueError:
                                current_date = date_text  # fallback to raw string
                        print(f"\n  Date: {current_date}")
                except NoSuchElementException:
                    pass
                continue

            if "Opta-prematch" in classes:
                print(f"  [{i}] First unplayed match — stopping.")
                break

            if "Opta-result" not in classes:
                continue

            record = {"date": current_date, "home_team": None, "away_team": None, "attendance": None}

            try:
                score_row = tbody.find_element(By.CSS_SELECTOR, "tr.Opta-Scoreline")
                record["home_team"] = get_team_name(score_row, "Home")
                record["away_team"] = get_team_name(score_row, "Away")

                button = score_row.find_element(By.CSS_SELECTOR, "button.Opta-Nest-Control")
                expansion_id = button.get_attribute("data-expansion_id")

                driver.execute_script("arguments[0].click();", button)

                try:
                    wait.until(EC.presence_of_element_located((By.XPATH,
                        f"//*[@id='{expansion_id}']//div[@class='Opta-Matchdata']//dt[text()='Attendance']"
                    )))
                    panel = driver.find_element(By.ID, expansion_id)
                    att_el = panel.find_element(
                        By.XPATH, ".//div[@class='Opta-Matchdata']//dt[text()='Attendance']/following-sibling::dd[1]"
                    )
                    record["attendance"] = att_el.get_attribute('innerHTML').strip() or None
                except (NoSuchElementException, TimeoutException):
                    record["attendance"] = None

                try:
                    all_rows2 = driver.find_elements(By.CSS_SELECTOR, "table tbody")
                    tbody2 = all_rows2[i]
                    score_row2 = tbody2.find_element(By.CSS_SELECTOR, "tr.Opta-Scoreline")
                    button2 = score_row2.find_element(By.CSS_SELECTOR, "button.Opta-Nest-Control")
                    driver.execute_script("arguments[0].click();", button2)
                except Exception:
                    pass
                time.sleep(0.3)

                print(f"  [{i}] {record['date']} | {record['home_team']} vs {record['away_team']} | Att: {record['attendance']}")

            except Exception as e:
                print(f"  [{i}] Error: {e}")

            records.append(record)

    finally:
        driver.quit()

    return pd.DataFrame(records)


if __name__ == "__main__":
    df = scrape_schedule(URL)
    print(f"\n{df.head(10)}")

    output_file = "L1_attendance_data.xlsx"
    df.to_excel(output_file, index=False)

    # Apply m/d date format to the date column so Excel displays 6/13 and sorts correctly
    wb = load_workbook(output_file)
    ws = wb.active
    for cell in ws["A"][1:]:  # skip header row
        if isinstance(cell.value, datetime):
            cell.number_format = "m/d"
    wb.save(output_file)